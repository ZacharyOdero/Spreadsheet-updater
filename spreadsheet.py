#this program takes value of a row in a spreedsheet, performs a certain operation and generates a new file with updated values
#importing relevant modules to update spreadsheets
import openpyxl as xl 
from openpyxl.chart import BarChart, Reference

#defining a function to update spreadsheets and generate a graph
def process_workbook(filename, sheet_name, cell_reference, file_save):
#updating spreadsheet
    wb = xl.load_workbook(filename) #creating an object of the xl class and saving it as a workbook
    sheet = wb[sheet_name] 
    for row in range(2, sheet.max_row + 1): #we add plus 1 as the max_row function takes values up to the second last values
        cell = sheet.cell(row, 7) #defining cell with targeted value
        new_price = cell.value * 0.9
        new_price_cell = sheet.cell(row, 17) #defining cell to store the updated value
        new_price_cell.value = new_price

#generating a graph
    values = Reference(sheet,
                       min_row=2,
                       max_row=sheet.max_row,
                       min_col=17,
                       max_col=17)
    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, cell_reference)
    wb.save(file_save) #saving the updated file. file_save = name of the updated file.
